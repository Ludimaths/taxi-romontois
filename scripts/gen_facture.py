#!/usr/bin/env python3
"""
gen_facture.py — Génère la facture DGEO en Excel avec openpyxl.
Lit les données JSON depuis stdin, écrit les bytes xlsx sur stdout.
"""

import sys
import json
import io
from datetime import date
from calendar import monthrange

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MOIS_NOMS = [
    "", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
]

# Couleurs (ARGB — openpyxl)
NAVY  = "FF0D3B7A"
GRAY  = "FFE2E8F0"
WHITE = "FFFFFFFF"

fill_navy = PatternFill("solid", fgColor=NAVY)
fill_gray = PatternFill("solid", fgColor=GRAY)

thin   = Side(border_style="thin")
medium = Side(border_style="medium")

font_title      = Font(bold=True, size=18)
font_annexe     = Font(bold=True, size=14)
font_section    = Font(bold=True, size=14, color=WHITE)
font_header     = Font(bold=True, size=11, color=WHITE)
font_lbl_bold   = Font(bold=True, size=11)
font_lbl        = Font(size=11)
font_data       = Font(size=11)
font_data_bold  = Font(bold=True, size=11)
font_total_gray = Font(bold=True, size=12)
font_tva        = Font(size=12)
font_total_navy = Font(bold=True, size=12, color=WHITE)
font_payment    = Font(size=14)

align_c  = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_l  = Alignment(horizontal="left",   vertical="center")
align_r  = Alignment(horizontal="right",  vertical="center")
align_v  = Alignment(vertical="center")
align_vw = Alignment(vertical="center", wrap_text=True)


def bdr(top_med=False, bot_med=False, left_med=False, right_med=False):
    return Border(
        top    = medium if top_med   else thin,
        bottom = medium if bot_med   else thin,
        left   = medium if left_med  else thin,
        right  = medium if right_med else thin,
    )


def sc(ws, row, col, value=None, font=None, fill=None, border=None,
       alignment=None, num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:       c.font      = font
    if fill:       c.fill      = fill
    if border:     c.border    = border
    if alignment:  c.alignment = alignment
    if num_format: c.number_format = num_format
    return c


def jours_ouvres_du_mois(annee, mois):
    _, nb = monthrange(annee, mois)
    return [
        date(annee, mois, d)
        for d in range(1, nb + 1)
        if date(annee, mois, d).weekday() < 5
    ]


def calcul(data):
    ecole    = data["ecole"]
    tournees = data["tournees"]
    eleves   = data["eleves"]
    mois     = data["mois"]
    annee    = data["annee"]
    ecole_id = ecole.get("id")

    jours = jours_ouvres_du_mois(annee, mois)
    tournees_ecole = [
        t for t in tournees
        if t.get("actif") and t.get("ecole_id") == ecole_id
    ]

    lignes = []
    for t in tournees_ecole:
        jour_sem   = t.get("jour_semaine", 0)
        circuit_id = t.get("circuit_id")
        km         = float(t.get("km") or 0)
        duree_min  = float(t.get("duree_minutes") or 0)
        prix_km    = float(t.get("prix_km") or 0)
        prix_heure = float(t.get("prix_heure") or 0)

        # python weekday: 0=lun…6=dim → jour_semaine: 1=lun…7=dim
        nb_tournees = sum(1 for d in jours if (d.weekday() + 1) == jour_sem)

        all_actifs   = [e for e in eleves if e.get("circuit_id") == circuit_id and e.get("actif")]
        ecole_actifs = [e for e in all_actifs if e.get("ecole_id") == ecole_id]
        total_el     = len(all_actifs)
        nb_ecole     = len(ecole_actifs)

        cout_tournee = km * prix_km + (duree_min / 60) * prix_heure
        cout_ecole = (
            round((cout_tournee / total_el) * nb_ecole * nb_tournees * 100) / 100
            if total_el > 0 else 0.0
        )

        lignes.append({
            "nom":          t.get("nom", ""),
            "nb_tournees":  nb_tournees,
            "km":           km,
            "duree_min":    int(duree_min),
            "cout_tournee": round(cout_tournee * 100) / 100,
            "total_eleves": total_el,
            "nb_ecole":     nb_ecole,
            "cout_ecole":   cout_ecole,
        })

    total_ht  = round(sum(l["cout_ecole"] for l in lignes) * 100) / 100
    tva_val   = round(total_ht * 0.081 * 100) / 100
    total_ttc = round((total_ht + tva_val) * 100) / 100
    prix_km_r = float(tournees_ecole[0].get("prix_km", 0)) if tournees_ecole else 0.0
    prix_hr_r = float(tournees_ecole[0].get("prix_heure", 0)) if tournees_ecole else 0.0

    return lignes, total_ht, tva_val, total_ttc, prix_km_r, prix_hr_r


def build_excel(data):
    params  = data.get("params", {})
    ecole   = data.get("ecole", {})
    mois    = data.get("mois", 1)
    annee   = data.get("annee", 2026)
    num_fac = data.get("numFacture", "")

    lignes, total_ht, tva_val, total_ttc, prix_km_r, prix_hr_r = calcul(data)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = " 6b) Facture - Exemple"

    # Largeurs colonnes (du modèle officiel)
    for i, w in enumerate([44.26, 13.59, 10.76, 10.76, 12.51, 12.51, 16.54, 22.06], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Hauteurs lignes (du modèle officiel)
    ws.row_dimensions[1].height = 18.75
    for r in range(2, 14):
        ws.row_dimensions[r].height = 20.25
    ws.row_dimensions[14].height = 7.5
    ws.row_dimensions[15].height = 19.5
    ws.row_dimensions[16].height = 66.95
    for r in range(17, 44):
        ws.row_dimensions[r].height = 15.0
    ws.row_dimensions[44].height = 20.1
    ws.row_dimensions[45].height = 20.1
    ws.row_dimensions[46].height = 15.75
    ws.row_dimensions[47].height = 18.0

    # ── R1 : Nom entreprise + ANNEXE 5b ──────────────────────────────────────
    nom_ent = params.get("nom") or "Nom ou logo de l'entreprise"
    sc(ws, 1, 1, " " + nom_ent, font=font_title, alignment=align_v)
    sc(ws, 1, 8, "ANNEXE 5b",   font=font_annexe, alignment=align_r)

    # ── R2-R5 : Infos entreprise (colonne A + B:C) ────────────────────────────
    sc(ws, 2, 1, "Adresse :",   font=font_lbl_bold, alignment=align_v)
    sc(ws, 2, 2, params.get("adresse") or "", font=font_lbl, alignment=align_v)
    sc(ws, 3, 1, "Téléphone :", font=font_lbl_bold, alignment=align_v)
    sc(ws, 3, 2, params.get("telephone") or "", font=font_lbl, alignment=align_v)
    sc(ws, 4, 1, "N° TVA :",    font=font_lbl_bold, alignment=align_v)
    sc(ws, 4, 2, params.get("tva") or "", font=font_lbl, alignment=align_v)
    sc(ws, 5, 1, "IBAN:",       font=font_lbl_bold, alignment=align_v)
    sc(ws, 5, 2, params.get("iban") or "", font=font_lbl, alignment=align_v)

    # ── R3-R5 : Infos école (colonne F + G:H) ────────────────────────────────
    sc(ws, 3, 6, "Nom de l'établissement/structure",   font=font_lbl_bold, alignment=align_vw)
    sc(ws, 3, 7, ecole.get("nom") or "",               font=font_lbl,      alignment=align_v)
    sc(ws, 4, 6, "Nom et prénom (Resp. facturation)",  font=font_lbl_bold, alignment=align_vw)
    sc(ws, 4, 7, ecole.get("nom_responsable_facturation") or "", font=font_lbl, alignment=align_v)
    sc(ws, 5, 6, "Adresse",                            font=font_lbl_bold, alignment=align_v)
    sc(ws, 5, 7, ecole.get("adresse") or "",           font=font_lbl,      alignment=align_v)

    # ── R7-R13 : Infos facture ────────────────────────────────────────────────
    sc(ws, 7,  1, "Établissement :",                   font=font_lbl_bold, alignment=align_v)
    sc(ws, 7,  2, ecole.get("nom") or "",              font=font_lbl,      alignment=align_v)
    sc(ws, 8,  1, "Structure(s) de l'établissement :", font=font_lbl_bold, alignment=align_v)
    sc(ws, 8,  2, "",                                  font=font_lbl,      alignment=align_v)
    sc(ws, 9,  1, "Facture N° :",                      font=font_lbl_bold, alignment=align_v)
    sc(ws, 9,  2, num_fac,                             font=font_data_bold, alignment=align_v)
    sc(ws, 10, 1, "Mois / année :",                    font=font_lbl,      alignment=align_v)
    sc(ws, 10, 2, f"{MOIS_NOMS[mois]} {annee}",        font=font_lbl,      alignment=align_v)
    sc(ws, 11, 1, "Lot :",                             font=font_lbl,      alignment=align_v)
    sc(ws, 11, 2, ecole.get("lot") or "",              font=font_lbl,      alignment=align_v)
    sc(ws, 12, 1, "Prix/km (hors TVA) :",              font=font_lbl,      alignment=align_v)
    sc(ws, 12, 2, prix_km_r,   font=font_lbl, alignment=align_r, num_format="0.00")
    sc(ws, 13, 1, "Prix/heure (hors TVA) :",           font=font_lbl,      alignment=align_v)
    sc(ws, 13, 2, prix_hr_r,   font=font_lbl, alignment=align_r, num_format="0.00")

    # ── Fusions (du modèle officiel) ──────────────────────────────────────────
    for rng in [
        "B2:C2", "B3:C3", "G3:H3",
        "B4:C4", "G4:H4",
        "B5:C5", "G5:H5",
        "B7:C7", "B8:C8",
        "B9:C9", "B10:C10", "B11:C11", "B12:C12", "B13:C13",
        "A15:H15",
    ]:
        ws.merge_cells(rng)

    # ── R15 : "Transports scolaires" — navy, borders medium ──────────────────
    sc(ws, 15, 1, "Transports scolaires",
       font=font_section, fill=fill_navy,
       border=Border(top=medium, bottom=medium, left=medium, right=thin),
       alignment=align_c)
    for col in range(2, 8):
        sc(ws, 15, col, None, fill=fill_navy,
           border=Border(top=medium, bottom=medium, left=thin, right=thin))
    sc(ws, 15, 8, None, fill=fill_navy,
       border=Border(top=medium, bottom=medium, left=thin, right=medium))

    # ── R16 : En-têtes colonnes — navy ───────────────────────────────────────
    headers = [
        "Nom de la tournée",
        "Nb. de\ntournées",
        "Distance\n(km)",
        "Durée\n(min)",
        "Coût tournée\n(hors TVA)",
        "Nb. total\nélèves",
        "Nb. élèves\nécole",
        "Coût école\n(hors TVA)",
    ]
    for i, h in enumerate(headers, 1):
        sc(ws, 16, i, h,
           font=font_header, fill=fill_navy,
           border=bdr(left_med=(i == 1), right_med=(i == 8)),
           alignment=align_c)

    # ── R17-R43 : Données ─────────────────────────────────────────────────────
    for idx, ligne in enumerate(lignes):
        r = 17 + idx
        sc(ws, r, 1, ligne["nom"],          font=font_data,      border=bdr(left_med=True), alignment=align_l)
        sc(ws, r, 2, ligne["nb_tournees"],  font=font_data,      border=bdr(),              alignment=align_c, num_format="0")
        sc(ws, r, 3, ligne["km"],           font=font_data,      border=bdr(),              alignment=align_c, num_format="0.00")
        sc(ws, r, 4, ligne["duree_min"],    font=font_data,      border=bdr(),              alignment=align_c, num_format="0")
        sc(ws, r, 5, ligne["cout_tournee"], font=font_data,      border=bdr(),              alignment=align_c, num_format="0.00")
        sc(ws, r, 6, ligne["total_eleves"], font=font_data,      border=bdr(),              alignment=align_c, num_format="0")
        sc(ws, r, 7, ligne["nb_ecole"],     font=font_data,      border=bdr(),              alignment=align_c, num_format="0")
        sc(ws, r, 8, ligne["cout_ecole"],   font=font_data_bold, border=bdr(right_med=True), alignment=align_c, num_format="0.00")

    # Lignes vides (bordures uniquement)
    for r in range(17 + len(lignes), 44):
        sc(ws, r, 1, None, border=bdr(left_med=True))
        for col in range(2, 8):
            sc(ws, r, col, None, border=bdr())
        sc(ws, r, 8, None, border=bdr(right_med=True))

    # ── R44 : Total HT — gris ────────────────────────────────────────────────
    sc(ws, 44, 1, "Total (sans TVA)", font=font_total_gray, fill=fill_gray,
       border=bdr(left_med=True), alignment=align_l)
    for col in range(2, 8):
        sc(ws, 44, col, None, fill=fill_gray, border=bdr())
    sc(ws, 44, 8, total_ht, font=font_total_gray, fill=fill_gray,
       border=bdr(right_med=True), alignment=align_c, num_format="0.00")

    # ── R45 : TVA 8.1% — gris ────────────────────────────────────────────────
    sc(ws, 45, 1, "TVA 8.1%", font=font_tva, fill=fill_gray,
       border=bdr(left_med=True), alignment=align_l)
    for col in range(2, 8):
        sc(ws, 45, col, None, fill=fill_gray, border=bdr())
    sc(ws, 45, 8, tva_val, font=font_tva, fill=fill_gray,
       border=bdr(right_med=True), alignment=align_c, num_format="0.00")

    # ── R46 : Total TTC — navy ───────────────────────────────────────────────
    sc(ws, 46, 1, "Total (avec TVA)", font=font_total_navy, fill=fill_navy,
       border=bdr(left_med=True, bot_med=True), alignment=align_l)
    for col in range(2, 8):
        sc(ws, 46, col, None, fill=fill_navy, border=bdr(bot_med=True))
    sc(ws, 46, 8, total_ttc, font=font_total_navy, fill=fill_navy,
       border=bdr(right_med=True, bot_med=True), alignment=align_c, num_format="0.00")

    # ── R47 : Paiement à 30 jours ────────────────────────────────────────────
    sc(ws, 47, 1, "Paiement à 30 jours", font=font_payment, alignment=align_v)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def main():
    raw  = sys.stdin.buffer.read()
    data = json.loads(raw.decode("utf-8-sig"))
    xlsx = build_excel(data)
    sys.stdout.buffer.write(xlsx)


if __name__ == "__main__":
    main()
